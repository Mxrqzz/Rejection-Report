from openpyxl import load_workbook, Workbook

planilha = load_workbook("planilha/data.xlsx")

dados = planilha["Analise Dezembro"]

# Dicionarios para contar numero de assunto e numero de assunto indeferidos
assunto_count = {}

assunto_count_indeferidas = {}

# Conta os assunto
for row in range(6, dados.max_row + 1):

    assunto = dados.cell(row=row, column=5).value

    if assunto not in assunto_count:

        assunto_count[assunto] = 0

    assunto_count[assunto] += 1

    # Verifica se a situação é "Indeferida"

    situacao = dados.cell(row=row, column=4).value

    if situacao is not None and situacao.strip().lower() == "não anuído":

        if assunto not in assunto_count_indeferidas:

            assunto_count_indeferidas[assunto] = 0

        assunto_count_indeferidas[assunto] += 1

print("assuntos Mapeadas...")

print(len(assunto_count))

# Criando nova planilha

analise_dados = Workbook()

analise = analise_dados.active

analise.title = "Avaliacao Critica Dezembro"

# Definindo Cabeçalho

analise["A1"] = "assuntos"

analise["B1"] = "Total de Anuidos"

analise["C1"] = "Total Analisados"

# Inserindo dados na planilha

for row, assunto in enumerate(assunto_count, start=2):

    analise[f"A{row}"] = assunto

    analise[f"B{row}"] = assunto_count_indeferidas.get(assunto, 0)

    analise[f"C{row}"] = assunto_count.get(assunto, 0)


analise_dados.save(filename="planilhas/dadosRelatorio.xlsx")

print("Planilha salva com sucesso...")
