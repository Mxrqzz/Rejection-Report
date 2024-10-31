from openpyxl import load_workbook, Workbook

planilha = load_workbook("planilha/data.xlsx")

dados = planilha["Dados Dezembro"]

# Criando Dicionarios que vao guardar a contagem
ncm_count = {}

ncm_count_indeferidas = {}

erroDePeticao_count = {}

ausenciaDocumento_count = {}

produtoIrregular_count = {}

outros_count = {}

# Contando numero de ncm
for row in range(6, dados.max_row + 1):

    ncm = dados.cell(row=row, column=9).value

    if ncm not in ncm_count:

        ncm_count[ncm] = 0

    ncm_count[ncm] += 1

    # Verifica se a situação é "Indeferida"

    situacao = dados.cell(row=row, column=13).value

    if situacao is not None and situacao.strip().lower() == "indeferida":

        if ncm not in ncm_count_indeferidas:

            ncm_count_indeferidas[ncm] = 0

        ncm_count_indeferidas[ncm] += 1

        # Verifica o resumo de diagnóstico

        resumo_diagnostico = dados.cell(row=row, column=15).value

        if resumo_diagnostico:

            resumo_diagnostico = resumo_diagnostico.strip().lower()

            if resumo_diagnostico == "erro de petição/código/informações":

                if ncm not in erroDePeticao_count:

                    erroDePeticao_count[ncm] = 0

                erroDePeticao_count[ncm] += 1

            elif resumo_diagnostico == "ausência de documento obrigatório":

                if ncm not in ausenciaDocumento_count:

                    ausenciaDocumento_count[ncm] = 0

                ausenciaDocumento_count[ncm] += 1

            elif resumo_diagnostico == "produto irregular":

                if ncm not in produtoIrregular_count:

                    produtoIrregular_count[ncm] = 0

                produtoIrregular_count[ncm] += 1

            elif resumo_diagnostico == "outro(s)":

                if ncm not in outros_count:

                    outros_count[ncm] = 0

                outros_count[ncm] += 1

print("NCMs Mapeadas...")

# Criando nova planilha
relatorio_dados = Workbook()

analise = relatorio_dados.active

analise.title = "Analise Dezembro"

# Definindo Cabeçalho
analise["A1"] = "NCMs"

analise["B1"] = "Erro de petição/código/informações"

analise["C1"] = "Ausência de documento obrigatório"

analise["D1"] = "Empresa irregular"

analise["E1"] = "Produto irregular"

analise["F1"] = "Produto e Empresa irregulares"

analise["G1"] = "Outros"

analise["H1"] = "Total de Indeferidos"

analise["I1"] = "Termo de Interdição"

analise["J1"] = "Total Analisados"


# Inserindo dados na planilha

for row, ncm in enumerate(ncm_count, start=2):

    analise[f"A{row}"] = ncm

    analise[f"B{row}"] = erroDePeticao_count.get(ncm, 0)

    analise[f"C{row}"] = ausenciaDocumento_count.get(ncm, 0)

    analise[f"E{row}"] = produtoIrregular_count.get(ncm, 0)

    analise[f"G{row}"] = outros_count.get(ncm, 0)

    analise[f"H{row}"] = ncm_count_indeferidas.get(ncm, 0)

    analise[f"J{row}"] = ncm_count.get(ncm, 0)


relatorio_dados.save(
    filename="C:/Users/rodrigo.filho/Desktop/mxrqzz/excel automation/planilhas/relatorio Dezembro.xlsx"
)

print("Planilha salva com sucesso...")
